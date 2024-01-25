import openpyxl
import os
from PIL import Image
from io import BytesIO
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive 
from openpyxl.utils import get_column_letter, column_index_from_string

contest_entry_img_path = r'C:\Users\Risinu Wijesinghe\OneDrive\Desktop\Projects\Frame_Generator\contest_entry_imgs'
participant_img_path = r'C:\Users\Risinu Wijesinghe\OneDrive\Desktop\Projects\Frame_Generator\participant_imgs'

gauth = GoogleAuth()
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth)

contest_entry_img_column = 'L'
participant_img_column = 'J'
caption_column =column_index_from_string('R')
name_column =column_index_from_string('C')

def get_photos(path,count):

    count = count
    excel_path =path
    work_book = openpyxl.load_workbook(excel_path)
    sheet = work_book.active

    err_numbers = []
    iteration_count = 0

    for img_cell, participant_img_cell in zip(sheet[contest_entry_img_column], sheet[participant_img_column]):
        if isinstance(img_cell.value, str) and 'drive.google.com' in img_cell.value:
            file_id = img_cell.value.split('id=')[-1]
            participant_file_id = participant_img_cell.value.split('id=')[-1]
            
            file = drive.CreateFile({'id':file_id})
            participant_file = drive.CreateFile({'id':participant_file_id})

            file.GetContentFile(os.path.join(contest_entry_img_path, f'{iteration_count} image.jpg'))
            participant_file.GetContentFile(os.path.join(participant_img_path, f'{iteration_count} participant_img.jpg'))

            iteration_count += 1

            if iteration_count >= count + len(err_numbers):
                break
        else:
            err_numbers.append(iteration_count)
            iteration_count += 1    
            continue

    
    contest_img_paths = [os.path.join(contest_entry_img_path, file) for file in os.listdir(contest_entry_img_path)]
    contest_images = [Image.open(img_path) for img_path in contest_img_paths]

    participant_img_paths = [os.path.join(participant_img_path, file) for file in os.listdir(participant_img_path)]
    participant_images = [Image.open(img_path) for img_path in participant_img_paths]


    if len(err_numbers) > 0:
        err_numbers_str = ', '.join(map(str, err_numbers))
        print(f'Photos {err_numbers_str} were not downloaded.')

    work_book.close()

    return contest_images,participant_images,err_numbers

def get_names_and_captions(path,count,err_nums=[]):
    names=[]
    captions=[]
    excel_path =path
    work_book = openpyxl.load_workbook(excel_path)
    sheet = work_book.active

    for i in range(2,count +len(err_nums)):
        names.append(sheet.cell(i,name_column).value)
        captions.append(sheet.cell(i,caption_column).value)

    return names,captions





